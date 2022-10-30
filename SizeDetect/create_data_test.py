from openpyxl import Workbook
from find_feature import find_features
from preprocessing import preprocessing_img
import cv2

wb = Workbook()
sheet1 = wb.active
sheet1.title = "size16"
sheet2 = wb.create_sheet(title="size18")
sheet3 = wb.create_sheet(title="size20")
sheet4 = wb.create_sheet(title="error_size16")
sheet5 = wb.create_sheet(title="error_size18")
sheet6 = wb.create_sheet(title="error_size20")

sheet1['A1'] = 'd1'
sheet1['B1'] = 'd2'
sheet1['C1'] = 'd3'
sheet1['D1'] = 'y'
sheet2['A1'] = 'd1'
sheet2['B1'] = 'd2'
sheet2['C1'] = 'd3'
sheet2['D1'] = 'y'
sheet3['A1'] = 'd1'
sheet3['B1'] = 'd2'
sheet3['C1'] = 'd3'
sheet3['D1'] = 'y'
sheet4['A1'] = 'd1'
sheet4['B1'] = 'd2'
sheet4['C1'] = 'd3'
sheet4['D1'] = 'y'
sheet5['A1'] = 'd1'
sheet5['B1'] = 'd2'
sheet5['C1'] = 'd3'
sheet5['D1'] = 'y'
sheet6['A1'] = 'd1'
sheet6['B1'] = 'd2'
sheet6['C1'] = 'd3'
sheet6['D1'] = 'y'


def create_workbook(path, index, size, d_1, d_2, d_3):
    col1name = 'A'
    col2name = 'B'
    col3name = 'C'
    col4name = 'D'
    col1name = col1name + str(index)
    col2name = col2name + str(index)
    col3name = col3name + str(index)
    col4name = col4name + str(index)
    if size == 16:
        sheet1[col1name] = d_1
        sheet1[col2name] = d_2
        sheet1[col3name] = d_3
        sheet1[col4name] = 1
    elif size == 18:
        sheet2[col1name] = d_1
        sheet2[col2name] = d_2
        sheet2[col3name] = d_3
        sheet2[col4name] = 1
    elif size == 20:
        sheet3[col1name] = d_1
        sheet3[col2name] = d_2
        sheet3[col3name] = d_3
        sheet3[col4name] = 1
    elif size == 0:
        sheet4[col1name] = d_1
        sheet4[col2name] = d_2
        sheet4[col3name] = d_3
        sheet4[col4name] = 0
    elif size == 1:
        sheet5[col1name] = d_1
        sheet5[col2name] = d_2
        sheet5[col3name] = d_3
        sheet5[col4name] = 0
    elif size == 2:
        sheet6[col1name] = d_1
        sheet6[col2name] = d_2
        sheet6[col3name] = d_3
        sheet6[col4name] = 0
    wb.save(path)


# size 16
txt16 = 'test image/size16/dep16 ({0}).jpg'
for i in range(6):
    img = cv2.imread(txt16.format(i + 1))
    img_org, img_contour = preprocessing_img(img)
    image, center, d1, d2, d3 = find_features(img_org, img_contour)
    create_workbook("test.xlsx", i + 2, 16, d1, d2, d3)

# size 18
txt18 = 'test image/size18/dep18 ({0}).jpg'
for i in range(6):
    img = cv2.imread(txt18.format(i + 1))
    img_org, img_contour = preprocessing_img(img)
    image, center, d1, d2, d3 = find_features(img_org, img_contour)
    create_workbook("test.xlsx", i + 2, 18, d1, d2, d3)

# size 20
txt20 = 'test image/size20/dep20 ({0}).jpg'
for i in range(6):
    img = cv2.imread(txt20.format(i + 1))
    img_org, img_contour = preprocessing_img(img)
    image, center, d1, d2, d3 = find_features(img_org, img_contour)
    create_workbook("test.xlsx", i + 2, 20, d1, d2, d3)

# not size 16
txt_error16 = 'test image/error/error ({0}).jpg'
for i in range(12):
    img = cv2.imread(txt_error16.format(i + 1))
    img_org, img_contour = preprocessing_img(img)
    image, center, d1, d2, d3 = find_features(img_org, img_contour)
    create_workbook("test.xlsx", i + 2, 0, d1, d2, d3)

# not size 18
txt_error18 = 'test image/error/error ({0}).jpg'
for i in range(12):
    img = cv2.imread(txt_error18.format(i + 1))
    img_org, img_contour = preprocessing_img(img)
    image, center, d1, d2, d3 = find_features(img_org, img_contour)
    create_workbook("test.xlsx", i + 2, 1, d1, d2, d3)

# not size 20
txt_error20 = 'test image/error/error ({0}).jpg'
for i in range(12):
    img = cv2.imread(txt_error20.format(i + 1))
    img_org, img_contour = preprocessing_img(img)
    image, center, d1, d2, d3 = find_features(img_org, img_contour)
    create_workbook("test.xlsx", i + 2, 2, d1, d2, d3)
